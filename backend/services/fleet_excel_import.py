"""
Fleet Excel import helpers (preview + commit).

Rules:
- One sheet per company
- Columns include MATRICULA and PLAZAS (with aliases)
- PLAZAS values can be `N` or `N+X+Y` (base + PMR extras)
"""

from __future__ import annotations

import io
import re
import unicodedata
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from db import crud as db_crud
from services.fleet_repository import FleetRepository


PLATE_HEADERS = {"matricula", "plate", "placa", "matric"}
SEATS_HEADERS = {"plazas", "plaza", "capacidad", "seats"}
COMPANY_HEADERS = {"empresa", "company", "operador", "sociedad"}


def _slugify(value: str) -> str:
    base = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")
    return base or "company"


def _normalize_plate(value: Any) -> str:
    return str(value or "").strip().upper()


def _normalize_header_label(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    normalized = unicodedata.normalize("NFKD", raw)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", ascii_only).strip()


def _parse_seats(value: Any) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[str]]:
    raw = str(value or "").strip()
    if not raw:
        return None, None, None, "PLAZAS vacio"

    compact = raw.replace(" ", "")
    if re.fullmatch(r"\d+", compact):
        base = int(compact)
        if base <= 0:
            return None, None, None, "PLAZAS invalido"
        return base, 0, base, None

    if re.fullmatch(r"\d+(?:\+\d+)+", compact):
        parts = [int(part) for part in compact.split("+")]
        if not parts or parts[0] <= 0:
            return None, None, None, "PLAZAS invalido"
        base = parts[0]
        pmr = sum(parts[1:])
        total = base + pmr
        return base, pmr, total, None

    return None, None, None, f"PLAZAS no reconocido: '{raw}'"


def _detect_header_row(
    ws,
    max_scan_rows: int = 8,
) -> Tuple[Optional[int], Dict[str, int], Dict[str, str]]:
    max_col = int(ws.max_column or 0)
    for row_idx in range(1, min(int(ws.max_row or 0), max_scan_rows) + 1):
        labels: Dict[str, int] = {}
        raw_labels: Dict[int, str] = {}
        for col_idx in range(1, max_col + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            label = _normalize_header_label(value)
            if not label:
                continue
            labels[label] = col_idx
            raw_labels[col_idx] = str(value or "").strip()

        plate_col = None
        seats_col = None
        company_col = None
        for label, col_idx in labels.items():
            if any(alias in label for alias in PLATE_HEADERS):
                plate_col = col_idx
            if any(alias in label for alias in SEATS_HEADERS):
                seats_col = col_idx
            if any(alias in label for alias in COMPANY_HEADERS):
                company_col = col_idx

        if plate_col and seats_col:
            columns = {"plate": plate_col, "seats": seats_col, "company": company_col or 1}
            normalized_columns = {
                "plate": raw_labels.get(plate_col, ""),
                "seats": raw_labels.get(seats_col, ""),
                "company": raw_labels.get(company_col or 1, ""),
            }
            return row_idx, columns, normalized_columns

    return None, {}, {}


def parse_fleet_excel_preview(file_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    sheets_payload: List[Dict[str, Any]] = []
    global_warnings: List[str] = []
    seen_global: set[Tuple[str, str]] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, columns, normalized_columns = _detect_header_row(ws)

        if header_row is None:
            sheets_payload.append(
                {
                    "sheet_name": sheet_name,
                    "company_name": sheet_name.strip() or sheet_name,
                    "company_slug": _slugify(sheet_name),
                    "header_detected": False,
                    "header_row": None,
                    "columns": {},
                    "normalized_columns": {},
                    "total_rows": 0,
                    "valid_rows": 0,
                    "invalid_rows": 0,
                    "warnings": ["No se detectaron columnas MATRICULA/PLAZAS"],
                    "vehicles": [],
                }
            )
            continue

        vehicles: List[Dict[str, Any]] = []
        warnings: List[str] = []
        valid_rows = 0
        invalid_rows = 0
        seen_sheet_plates: set[str] = set()

        for row_idx in range(header_row + 1, int(ws.max_row or 0) + 1):
            plate_raw = ws.cell(row=row_idx, column=columns["plate"]).value
            seats_raw = ws.cell(row=row_idx, column=columns["seats"]).value
            company_raw = ws.cell(row=row_idx, column=columns.get("company", 1)).value

            plate = _normalize_plate(plate_raw)
            if not plate and (seats_raw is None or str(seats_raw).strip() == ""):
                continue
            if not plate:
                invalid_rows += 1
                warnings.append(f"Fila {row_idx}: matricula vacia")
                continue

            seats_base, seats_pmr, seats_total, seats_error = _parse_seats(seats_raw)
            if seats_error:
                invalid_rows += 1
                warnings.append(f"Fila {row_idx}: {seats_error}")
                continue

            if plate in seen_sheet_plates:
                invalid_rows += 1
                warnings.append(f"Fila {row_idx}: matricula duplicada en hoja ({plate})")
                continue
            seen_sheet_plates.add(plate)

            company_name = str(company_raw or "").strip() or sheet_name.strip() or sheet_name
            company_slug = _slugify(sheet_name)
            global_key = (company_slug, plate)
            if global_key in seen_global:
                warnings.append(f"Fila {row_idx}: duplicado global ({plate})")
            seen_global.add(global_key)

            valid_rows += 1
            vehicles.append(
                {
                    "row": row_idx,
                    "company_name": company_name,
                    "company_slug": company_slug,
                    "sheet_name": sheet_name,
                    "plate": plate,
                    "vehicle_code": plate,
                    "seats_base": int(seats_base or 0),
                    "seats_pmr": int(seats_pmr or 0),
                    "seats_total": int(seats_total or 0),
                    "status": "active",
                }
            )

        sheet_payload = {
            "sheet_name": sheet_name,
            "company_name": sheet_name.strip() or sheet_name,
            "company_slug": _slugify(sheet_name),
            "header_detected": True,
            "header_row": header_row,
            "columns": columns,
            "normalized_columns": normalized_columns,
            "total_rows": valid_rows + invalid_rows,
            "valid_rows": valid_rows,
            "invalid_rows": invalid_rows,
            "warnings": warnings,
            "vehicles": vehicles,
        }
        sheets_payload.append(sheet_payload)
        if warnings:
            global_warnings.extend([f"{sheet_name}: {warning}" for warning in warnings])

    return {
        "sheets": sheets_payload,
        "companies_detected": [sheet.get("company_name") for sheet in sheets_payload],
        "sheet_names": list(wb.sheetnames),
        "warnings": global_warnings,
    }


def commit_fleet_excel_import(
    db,
    *,
    file_bytes: bytes,
    primary_sheet_name: str,
    ute_name: Optional[str] = None,
    repository: Optional[FleetRepository] = None,
) -> Dict[str, Any]:
    preview = parse_fleet_excel_preview(file_bytes)
    sheets = preview.get("sheets", []) if isinstance(preview.get("sheets"), list) else []
    if not sheets:
        raise ValueError("Excel sin hojas validas")

    by_name = {str(sheet.get("sheet_name", "")): sheet for sheet in sheets}
    primary_sheet_name_clean = str(primary_sheet_name or "").strip()
    primary_sheet = by_name.get(primary_sheet_name_clean)
    if primary_sheet is None:
        primary_sheet = next(
            (
                sheet for sheet in sheets
                if str(sheet.get("sheet_name", "")).strip().lower() == primary_sheet_name_clean.lower()
            ),
            None,
        )

    if primary_sheet is None:
        raise ValueError("La hoja principal seleccionada no existe en el Excel")
    if not bool(primary_sheet.get("header_detected", False)):
        raise ValueError("La hoja principal no tiene columnas validas MATRICULA/PLAZAS")

    valid_sheets = [sheet for sheet in sheets if bool(sheet.get("header_detected", False))]
    if not valid_sheets:
        raise ValueError("Excel sin hojas validas MATRICULA/PLAZAS")

    fleet_repo = repository or FleetRepository()
    summary_by_company: Dict[str, Dict[str, Any]] = {}
    company_ids: List[str] = []
    primary_company_id: Optional[str] = None

    # Pass 1: ensure companies and commit so repository DB sessions can resolve them.
    sheet_to_company: Dict[str, Dict[str, str]] = {}
    for sheet in valid_sheets:
        sheet_name = str(sheet.get("sheet_name", "") or "").strip()
        company_name = str(sheet.get("company_name", "") or "").strip() or sheet_name
        company_slug = _slugify(sheet_name)
        company = db_crud.ensure_company(
            db,
            name=company_name,
            preferred_id=f"company_{company_slug}",
        )
        company_id = str(company.id)
        sheet_to_company[sheet_name] = {
            "company_id": company_id,
            "company_name": company_name,
        }
        if company_id not in company_ids:
            company_ids.append(company_id)
        if sheet_name == str(primary_sheet.get("sheet_name", "")).strip():
            primary_company_id = company_id
    db.commit()

    # Pass 2: upsert vehicles by (company_id, plate).
    for sheet in valid_sheets:
        sheet_name = str(sheet.get("sheet_name", "") or "").strip()
        company_info = sheet_to_company.get(sheet_name, {})
        company_id = str(company_info.get("company_id") or "").strip()
        company_name = str(company_info.get("company_name") or "").strip() or sheet_name
        if not company_id:
            continue

        existing = fleet_repo.list_vehicles(company_id=company_id)
        existing_by_plate = {str(v.get("plate", "")).strip().upper(): v for v in existing}

        created = 0
        updated = 0
        invalid = int(sheet.get("invalid_rows", 0) or 0)
        vehicles = sheet.get("vehicles", []) if isinstance(sheet.get("vehicles"), list) else []
        for row in vehicles:
            plate = _normalize_plate(row.get("plate"))
            if not plate:
                continue

            seats_base = int(row.get("seats_base") or 0)
            seats_pmr = int(row.get("seats_pmr") or 0)
            seats_total = int(row.get("seats_total") or 0)
            payload = {
                "company_id": company_id,
                "vehicle_code": str(row.get("vehicle_code") or plate),
                "plate": plate,
                "brand": None,
                "model": None,
                "year": None,
                "seats_base": seats_base,
                "seats_pmr": seats_pmr,
                "seats_min": max(1, seats_base or seats_total),
                "seats_max": max(1, seats_total),
                "status": "active",
                "fuel_type": None,
                "accessibility": bool(seats_pmr > 0),
                "mileage_km": None,
                "notes": (
                    f"Importado Excel {sheet_name} | PMR={seats_pmr}"
                    if seats_pmr > 0
                    else f"Importado Excel {sheet_name}"
                ),
                "documents": [],
            }
            existing_row = existing_by_plate.get(plate)
            if existing_row is None:
                fleet_repo.create_vehicle(payload, company_id=company_id)
                created += 1
            else:
                fleet_repo.update_vehicle(str(existing_row.get("id")), payload, company_id=company_id)
                updated += 1

        summary_by_company[company_id] = {
            "company_id": company_id,
            "company_name": company_name,
            "sheet_name": sheet_name,
            "created": created,
            "updated": updated,
            "invalid": invalid,
            "kept_active_missing": True,
            "total_processed": created + updated + invalid,
        }

    if not primary_company_id:
        raise ValueError("No se pudo resolver la empresa principal")

    default_ute_name = f"UTE {str(primary_sheet.get('sheet_name', primary_sheet_name_clean)).strip()}"
    ute = db_crud.create_or_update_ute(
        db,
        ute_name=str(ute_name or "").strip() or default_ute_name,
        owner_company_id=primary_company_id,
        member_company_ids=company_ids,
    )
    db.commit()

    return {
        "primary_company_id": primary_company_id,
        "ute_id": str(ute.id),
        "ute_name": str(ute.name or ""),
        "scope_mode_suggested": "company",
        "summary_by_company": list(summary_by_company.values()),
        "companies_count": len(company_ids),
        "total_created": sum(int(item["created"]) for item in summary_by_company.values()),
        "total_updated": sum(int(item["updated"]) for item in summary_by_company.values()),
        "total_invalid": sum(int(item["invalid"]) for item in summary_by_company.values()),
    }

