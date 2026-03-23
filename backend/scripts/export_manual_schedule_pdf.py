"""
Generate the same schedule PDF outside the app from a manual service sheet.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime, time
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_ROOT = REPO_ROOT / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from parser import parse_routes  # noqa: E402
from pdf_service import generate_schedule_pdf  # noqa: E402
from router_service import get_real_travel_time  # noqa: E402


EARTH_RADIUS_KM = 6371.0
FALLBACK_SPEED_KMH = 50
DEADHEAD_BUFFER_MINUTES = 3
FUZZY_MATCH_THRESHOLD = 0.70

DAY_NAME_MAP = {
    "todas   lunes (2)": "Lunes",
    "todas   martes (2)": "Martes",
    "todas   mi-ju-vi": "Miercoles-Jueves-Viernes",
}

NAME_ALIASES = {
    ("ue3617", "matama - castrelos - ies coruxo"): (
        "matama - castrelos (rua dos canteiros 6) - ies coruxo"
    ),
    ("ue3619", "san xoan - cep igrexa-valadares & eei monte do alba"): (
        "san xoan - cep igrexa-valadares"
    ),
}


@dataclass(frozen=True)
class RouteCandidate:
    contract_key: str
    route_name_key: str
    route_type: str
    official_minutes: int
    route: Any


def _normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("saida", "salida")
    text = text.replace("nº", "n")
    text = text.replace("º", "o")
    text = re.sub(r"\s+", " ", text)
    return text


def _sheet_day_name(sheet_name: str) -> str:
    return DAY_NAME_MAP.get(_normalize_text(sheet_name), sheet_name.strip())


def _slugify(value: str) -> str:
    normalized = _normalize_text(value)
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return normalized or "schedule"


def _time_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.time().strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    text = str(value).strip().replace(";", ":").replace(".", ":")
    if not text:
        return ""
    for chunk in (text[:5], text):
        try:
            return datetime.strptime(chunk, "%H:%M").strftime("%H:%M")
        except ValueError:
            continue
    match = re.search(r"(\d{1,2})\D+(\d{1,2})", text)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        if 0 <= hours <= 23 and 0 <= minutes <= 59:
            return f"{hours:02d}:{minutes:02d}"
    return text


def _time_to_minutes(value: Any) -> int:
    text = _time_to_str(value)
    if not text:
        return 24 * 60 + 999
    try:
        parsed = datetime.strptime(text[:5], "%H:%M")
    except ValueError:
        return 24 * 60 + 999
    return parsed.hour * 60 + parsed.minute


def _route_type_from_sentido(value: Any) -> str:
    sentido_key = _normalize_text(value)
    return "entry" if "entrada" in sentido_key else "exit"


def _route_official_minutes(route: Any) -> int:
    raw = route.arrival_time if route.type == "entry" else route.departure_time
    return _time_to_minutes(raw)


def _stop_to_dict(stop: Any) -> Dict[str, Any]:
    if hasattr(stop, "model_dump"):
        return stop.model_dump()
    return stop.dict()


def _route_fingerprint(candidate: RouteCandidate) -> Tuple[Any, ...]:
    stops = tuple(
        (
            str(stop.name),
            round(float(stop.lat), 6),
            round(float(stop.lon), 6),
            int(stop.order),
        )
        for stop in candidate.route.stops
    )
    return (
        candidate.contract_key,
        candidate.route_name_key,
        candidate.route_type,
        candidate.official_minutes,
        stops,
    )


def _dedupe_candidates(candidates: Iterable[RouteCandidate]) -> List[RouteCandidate]:
    deduped: Dict[Tuple[Any, ...], RouteCandidate] = {}
    for candidate in candidates:
        deduped.setdefault(_route_fingerprint(candidate), candidate)
    return list(deduped.values())


def _load_catalog(routes_dir: Path) -> Tuple[List[RouteCandidate], Dict[str, Any]]:
    catalog: List[RouteCandidate] = []
    source_files = sorted(routes_dir.glob("UE36*.xlsx"))
    if not source_files:
        raise FileNotFoundError(f"No route catalog files found in {routes_dir}")

    for source_file in source_files:
        for route in parse_routes(str(source_file)):
            catalog.append(
                RouteCandidate(
                    contract_key=_normalize_text(route.contract_id),
                    route_name_key=_normalize_text(route.name),
                    route_type=str(route.type),
                    official_minutes=_route_official_minutes(route),
                    route=route,
                )
            )

    exact_index: Dict[Tuple[str, str, str], List[RouteCandidate]] = {}
    contract_type_index: Dict[Tuple[str, str], List[RouteCandidate]] = {}
    for candidate in catalog:
        exact_index.setdefault(
            (candidate.contract_key, candidate.route_name_key, candidate.route_type), []
        ).append(candidate)
        contract_type_index.setdefault(
            (candidate.contract_key, candidate.route_type), []
        ).append(candidate)

    return catalog, {
        "exact_index": exact_index,
        "contract_type_index": contract_type_index,
    }


def _apply_alias(contract_key: str, route_name_key: str) -> Tuple[str, bool]:
    alias = NAME_ALIASES.get((contract_key, route_name_key))
    if not alias:
        return route_name_key, False
    return alias, True


def _choose_best_candidate(
    candidates: Iterable[RouteCandidate],
    manual_minutes: int,
) -> RouteCandidate:
    deduped = _dedupe_candidates(candidates)
    if not deduped:
        raise LookupError("No route candidates available")
    return min(
        deduped,
        key=lambda candidate: (
            abs(candidate.official_minutes - manual_minutes),
            candidate.official_minutes,
            str(candidate.route.id),
        ),
    )


def _fuzzy_candidates(
    contract_type_index: Dict[Tuple[str, str], List[RouteCandidate]],
    contract_key: str,
    route_type: str,
    route_name_key: str,
) -> List[RouteCandidate]:
    pool = contract_type_index.get((contract_key, route_type), [])
    scored: List[Tuple[float, RouteCandidate]] = []
    for candidate in pool:
        ratio = SequenceMatcher(
            None,
            route_name_key,
            candidate.route_name_key,
        ).ratio()
        if ratio >= FUZZY_MATCH_THRESHOLD:
            scored.append((ratio, candidate))
    if not scored:
        return []
    best_ratio = max(score for score, _ in scored)
    return [candidate for score, candidate in scored if score == best_ratio]


def _match_route(
    indexes: Dict[str, Any],
    contract: Any,
    route_name: Any,
    route_type: str,
    manual_minutes: int,
) -> Tuple[RouteCandidate, str]:
    contract_key = _normalize_text(contract)
    route_name_key = _normalize_text(route_name)
    route_name_key, alias_used = _apply_alias(contract_key, route_name_key)

    exact_candidates = indexes["exact_index"].get(
        (contract_key, route_name_key, route_type),
        [],
    )
    if exact_candidates:
        method = "alias" if alias_used else "exact"
        return _choose_best_candidate(exact_candidates, manual_minutes), method

    fuzzy = _fuzzy_candidates(
        indexes["contract_type_index"],
        contract_key,
        route_type,
        route_name_key,
    )
    if fuzzy:
        return _choose_best_candidate(fuzzy, manual_minutes), "fuzzy"

    raise LookupError(
        f"Route not found for contract={contract!r}, name={route_name!r}, type={route_type!r}"
    )


def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    lat1_rad, lon1_rad, lat2_rad, lon2_rad = map(
        math.radians,
        (lat1, lon1, lat2, lon2),
    )
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2) ** 2
    )
    return 2 * EARTH_RADIUS_KM * math.asin(math.sqrt(a))


def _estimate_positioning_minutes(
    previous_item: Optional[Dict[str, Any]],
    current_item: Dict[str, Any],
) -> int:
    if not previous_item:
        return 0

    previous_stops = previous_item.get("stops") or []
    current_stops = current_item.get("stops") or []
    if not previous_stops or not current_stops:
        return 0

    previous_end = previous_stops[-1]
    current_start = current_stops[0]

    lat1 = float(previous_end.get("lat") or 0)
    lon1 = float(previous_end.get("lon") or 0)
    lat2 = float(current_start.get("lat") or 0)
    lon2 = float(current_start.get("lon") or 0)
    if not lat1 or not lon1 or not lat2 or not lon2:
        return 0

    osrm_minutes = get_real_travel_time(lat1, lon1, lat2, lon2)
    if osrm_minutes is not None:
        return int(osrm_minutes) + DEADHEAD_BUFFER_MINUTES

    km = _haversine_km(lat1, lon1, lat2, lon2)
    minutes = int((km / FALLBACK_SPEED_KMH) * 60)
    return max(5, minutes + DEADHEAD_BUFFER_MINUTES)


def _manual_row_to_item(
    indexes: Dict[str, Any],
    row: Tuple[Any, ...],
    stats: Dict[str, int],
) -> Dict[str, Any]:
    contract, route_name, sentido, start_raw, end_raw, *_rest = list(row) + [None] * 8
    route_type = _route_type_from_sentido(sentido)
    relevant_minutes = _time_to_minutes(end_raw if route_type == "entry" else start_raw)
    candidate, method = _match_route(
        indexes=indexes,
        contract=contract,
        route_name=route_name,
        route_type=route_type,
        manual_minutes=relevant_minutes,
    )
    stats[method] = stats.get(method, 0) + 1

    route = candidate.route
    stops = [_stop_to_dict(stop) for stop in route.stops]
    return {
        "route_id": str(route.id),
        "start_time": _time_to_str(start_raw),
        "end_time": _time_to_str(end_raw),
        "type": route_type,
        "capacity_needed": int(getattr(route, "capacity_needed", 0) or 0),
        "vehicle_capacity_min": getattr(route, "vehicle_capacity_min", None),
        "vehicle_capacity_max": getattr(route, "vehicle_capacity_max", None),
        "vehicle_capacity_range": getattr(route, "vehicle_capacity_range", None),
        "school_name": getattr(route, "school_name", None),
        "stops": stops,
        "contract_id": getattr(route, "contract_id", None),
        "manual_route_name": str(route_name or "").strip(),
    }


def _build_schedules(
    manual_excel: Path,
    indexes: Dict[str, Any],
) -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, int]]:
    workbook = load_workbook(manual_excel, data_only=True, read_only=True)
    schedules_by_sheet: Dict[str, List[Dict[str, Any]]] = {}
    stats: Dict[str, int] = {"exact": 0, "alias": 0, "fuzzy": 0}

    for sheet in workbook.worksheets:
        schedule: List[Dict[str, Any]] = []
        current_bus: Optional[Dict[str, Any]] = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            first_cell = row[0] if row else None
            if isinstance(first_cell, str) and first_cell.strip().upper().startswith("BUS "):
                current_bus = {"bus_id": first_cell.strip(), "items": []}
                schedule.append(current_bus)
                continue

            if not current_bus:
                continue

            contract = row[0] if len(row) > 0 else None
            route_name = row[1] if len(row) > 1 else None
            sentido = row[2] if len(row) > 2 else None
            if not contract or not route_name or not sentido:
                continue

            item = _manual_row_to_item(indexes=indexes, row=row, stats=stats)
            previous_item = current_bus["items"][-1] if current_bus["items"] else None
            positioning = _estimate_positioning_minutes(previous_item, item)
            item["positioning_minutes"] = positioning
            item["deadhead_minutes"] = positioning
            current_bus["items"].append(item)

        schedules_by_sheet[sheet.title] = [bus for bus in schedule if bus["items"]]

    return schedules_by_sheet, stats


def _write_pdf(output_path: Path, schedule: List[Dict[str, Any]], day_name: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pdf_buffer = generate_schedule_pdf(schedule, day_name=day_name)
    output_path.write_bytes(pdf_buffer.getvalue())


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export Tutti-style PDF from a manual service sheet.",
    )
    parser.add_argument(
        "manual_excel",
        type=Path,
        help="Path to the manual service workbook.",
    )
    parser.add_argument(
        "--routes-dir",
        type=Path,
        default=REPO_ROOT / "ejemplo excel rutas a optimizar",
        help="Directory with the source route catalogs.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=REPO_ROOT / "generated" / "manual_schedule_pdfs",
        help="Directory where the generated PDFs will be saved.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    _, indexes = _load_catalog(args.routes_dir)
    schedules_by_sheet, stats = _build_schedules(args.manual_excel, indexes)

    if not schedules_by_sheet:
        raise RuntimeError("No schedules were extracted from the manual workbook")

    print("Matched rows:", sum(stats.values()))
    for key in ("exact", "alias", "fuzzy"):
        print(f"  {key}: {stats.get(key, 0)}")

    for sheet_name, schedule in schedules_by_sheet.items():
        day_name = _sheet_day_name(sheet_name)
        filename = f"tutti_schedule_manual_{_slugify(day_name)}.pdf"
        output_path = args.output_dir / filename
        _write_pdf(output_path, schedule, day_name)
        print(f"PDF: {output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
